# Spreadsheet Cell Inverter

# Write a program to invert the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5
# (and vice versa). This should be done for all cells in the spreadsheet.

from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl as xl
import os, sys



os.chdir(r'D:\Python\Python_Practice\Practice')

wb = xl.load_workbook('example.xlsx')
sheet = wb.active

r_max = sheet.max_row
r_min = sheet.min_row

c_max = sheet.max_column
c_min = sheet.min_column

sheetData = []


for rowNum in range(r_min, r_max + 1):
    for col in range(c_min, c_max + 1):
        cell = sheet.cell(row=rowNum, column=col)
        
        dest_letter = get_column_letter(int(cell.coordinate[1:]))
        dest_digit = str(column_index_from_string(cell.coordinate[0]))
        dest_coord = dest_letter + dest_digit
                
        sheetData.append((dest_coord, cell.value))


sheet.delete_cols(1, c_max)

for coord, data in sheetData:
    sheet[coord] = data


wb.save('example_new.xlsx')



