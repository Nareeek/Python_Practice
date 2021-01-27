# Spreadsheet to Text Files
# The program should open a spreadsheet and write the cells of column A
# into one text file, the cells of column B into another text file, and so on.


from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl as xl
import os, sys


os.chdir(r'D:\Python\Python_Practice\Practice')

wb = xl.load_workbook('Text_Files_to_Spreadsheet.xlsx')
sheet = wb.active

new_files_prefix = 'Spreadsheet_to_text_files_'


c_min = sheet.min_column
c_max = sheet.max_column

def Int(string):
    import re
    pattern = r'\d+$'
    res = re.findall(pattern, string)
    res = ''.join(res)
    return int(res)
    

for col in range(c_min, c_max + 1):
    letter = get_column_letter(col)

    letter_max_row_number = len(sheet[letter])
    letter_min_row_number = Int(sheet[letter][0].coordinate)
    
    fileObj = open(new_files_prefix + letter + '_column.txt', 'w')

    for row in range(letter_min_row_number, letter_max_row_number):
        line = (sheet[letter + str(row)]).value
        fileObj.write(line)
        
    fileObj.close()
    
    print(col)
