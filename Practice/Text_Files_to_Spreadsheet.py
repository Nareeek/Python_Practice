# Text Files to Spreadsheet

# Write a program to read in the contents of several text files
# (you can make the text files yourself) and insert those contents
# into a spreadsheet, with one line of text per row. The lines of the
# first text file will be in the cells of column A, the lines of the second
# text file will be in the cells of column B, and so on.
#
# Use the readlines() File object method to return a list of strings,
# one string per line in the file. For the first file, output the first line
# to column 1, row 1. The second line should be written to column 1, row 2,
# and so on. The next file that is read with readlines() will be written to
# column 2, the next file to column 3, and so on.


from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl as xl
import os, sys


os.chdir(r'D:\Python\Python_Practice\Practice')

wb = xl.Workbook()
sheet = wb.active
order = 1

for files in os.listdir('.'):
    digit = 1
    if files.endswith('.txt') and files.startswith('Lorem'):
        print(files)
        fileObj = open(files, 'r')
        
        letter = get_column_letter(order)
        
        for line in fileObj.readlines():
            coord = letter + str(digit)
            sheet[coord] = line
            digit += 1
        order += 1
            

wb.save('Text_Files_to_Spreadsheet.xlsx')

